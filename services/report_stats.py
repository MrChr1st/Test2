import asyncio
import json
import logging
import uuid
import urllib.error
import urllib.request
from datetime import datetime, timedelta
from time_utils import MOSCOW_TZ, format_moscow, now_moscow_str
from pathlib import Path
from tempfile import NamedTemporaryFile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from report_settings import AUTO_REPORT_HOURS, PRIVATE_CHAT_ID, REPORT_BOT_TOKEN


EVENTS_FILE = "clientbot_events.jsonl"
STATE_FILE = "clientbot_auto_report_state.json"
_scheduler_started = False


def _now_str() -> str:
    return now_moscow_str()


def _append_event(event: dict) -> None:
    item = {"time": _now_str(), **event}
    with open(EVENTS_FILE, "a", encoding="utf-8") as f:
        f.write(json.dumps(item, ensure_ascii=False) + "\n")


def _read_events() -> list[dict]:
    path = Path(EVENTS_FILE)
    if not path.exists():
        return []
    items = []
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                items.append(json.loads(line))
            except Exception:
                continue
    return items


def _last_24h_events() -> list[dict]:
    cutoff = datetime.now(MOSCOW_TZ).replace(tzinfo=None) - timedelta(hours=24)
    result = []
    for item in _read_events():
        try:
            dt = datetime.strptime(item["time"], "%Y-%m-%d %H:%M:%S")
        except Exception:
            continue
        if dt >= cutoff:
            result.append(item)
    return result


def _post_json(url: str, payload: dict, timeout: int = 20):
    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=data,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        body = resp.read().decode("utf-8", errors="replace")
        return resp.status, body


def _build_multipart(fields: dict, file_field: str, filename: str, content: bytes, mime: str):
    boundary = "----WebKitFormBoundary" + uuid.uuid4().hex
    lines = []
    for key, value in fields.items():
        lines.append(f"--{boundary}\r\n".encode())
        lines.append(f'Content-Disposition: form-data; name="{key}"\r\n\r\n'.encode())
        lines.append(str(value).encode("utf-8"))
        lines.append(b"\r\n")
    lines.append(f"--{boundary}\r\n".encode())
    lines.append(
        f'Content-Disposition: form-data; name="{file_field}"; filename="{filename}"\r\n'.encode()
    )
    lines.append(f"Content-Type: {mime}\r\n\r\n".encode())
    lines.append(content)
    lines.append(b"\r\n")
    lines.append(f"--{boundary}--\r\n".encode())
    body = b"".join(lines)
    return body, boundary


def _post_multipart(url: str, fields: dict, file_field: str, filename: str, content: bytes, mime: str, timeout: int = 40):
    body, boundary = _build_multipart(fields, file_field, filename, content, mime)
    req = urllib.request.Request(
        url,
        data=body,
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return resp.status, resp.read().decode("utf-8", errors="replace")


def log_event_exchange_opened(user_id: int, username: str, profile_link: str) -> None:
    _append_event({
        "type": "opened",
        "user_id": user_id,
        "username": username,
        "profile_link": profile_link,
    })


def log_event_request_created(
    request_id: int,
    user_id: int,
    username: str,
    profile_link: str,
    from_currency: str,
    to_currency: str,
    amount_from: float,
    amount_to: float,
    receive_details: str,
    payment_method: str,
    payment_submethod: str,
) -> None:
    _append_event({
        "type": "request_created",
        "request_id": request_id,
        "user_id": user_id,
        "username": username,
        "profile_link": profile_link,
        "from_currency": from_currency,
        "to_currency": to_currency,
        "amount_from": amount_from,
        "amount_to": amount_to,
        "receive_details": receive_details,
        "payment_method": payment_method,
        "payment_submethod": payment_submethod,
    })


def log_event_paid(
    request_id: int,
    user_id: int,
    username: str,
    profile_link: str,
    from_currency: str,
    to_currency: str,
    amount_from: float,
    amount_to: float,
    receive_details: str,
    payment_method: str,
) -> None:
    _append_event({
        "type": "paid",
        "request_id": request_id,
        "user_id": user_id,
        "username": username,
        "profile_link": profile_link,
        "from_currency": from_currency,
        "to_currency": to_currency,
        "amount_from": amount_from,
        "amount_to": amount_to,
        "receive_details": receive_details,
        "payment_method": payment_method,
    })


def log_event_qr_requested(
    request_id: int,
    user_id: int,
    username: str,
    profile_link: str,
    asset_name: str,
    target_value: str,
) -> None:
    _append_event({
        "type": "qr_requested",
        "request_id": request_id,
        "user_id": user_id,
        "username": username,
        "profile_link": profile_link,
        "asset_name": asset_name,
        "target_value": target_value,
    })


def log_event_wallet_urgent(
    request_id: int,
    user_id: int,
    username: str,
    profile_link: str,
    from_currency: str,
    to_currency: str,
    amount_from: float,
    amount_to: float,
    receive_details: str,
) -> None:
    _append_event({
        "type": "wallet_urgent",
        "request_id": request_id,
        "user_id": user_id,
        "username": username,
        "profile_link": profile_link,
        "from_currency": from_currency,
        "to_currency": to_currency,
        "amount_from": amount_from,
        "amount_to": amount_to,
        "receive_details": receive_details,
    })


def _style_sheet(ws, title: str, headers: list[str]):
    ws.title = title
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 24
    return border


def _autosize(ws):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            val = "" if cell.value is None else str(cell.value)
            widths[cell.column] = max(widths.get(cell.column, 0), min(len(val) + 2, 40))
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _add_rows(ws, rows: list[list], border):
    alt_fill = PatternFill(fill_type="solid", fgColor="F7FBFF")
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if r_idx % 2 == 0:
                cell.fill = alt_fill


def _build_opened_rows(events: list[dict]) -> list[list]:
    return [
        [e["time"], e.get("username", ""), e.get("user_id", ""), e.get("profile_link", "")]
        for e in events if e.get("type") == "opened"
    ]


def _build_request_rows(events: list[dict]) -> list[list]:
    paid_map = {str(e.get("request_id", "")): e.get("time", "") for e in events if e.get("type") == "paid"}
    rows = []
    for e in events:
        if e.get("type") != "request_created":
            continue
        req = str(e.get("request_id", ""))
        rows.append([
            e["time"],
            req,
            e.get("username", ""),
            e.get("user_id", ""),
            e.get("profile_link", ""),
            e.get("from_currency", ""),
            e.get("to_currency", ""),
            e.get("amount_from", ""),
            e.get("amount_to", ""),
            e.get("receive_details", ""),
            f"{e.get('payment_method', '')} / {e.get('payment_submethod', '')}",
            "Оплачено" if req in paid_map else "Ожидает оплаты",
            paid_map.get(req, ""),
        ])
    return rows


def _financial_summary(events: list[dict]) -> dict:
    opened = sum(1 for e in events if e.get("type") == "opened")
    new_requests = [e for e in events if e.get("type") == "request_created"]
    paid_requests = [e for e in events if e.get("type") == "paid"]

    totals = {}
    for e in new_requests:
        cur = e.get("from_currency", "")
        amt = float(e.get("amount_from", 0) or 0)
        if cur:
            totals[cur] = totals.get(cur, 0.0) + amt

    return {
        "opened": opened,
        "new_requests": len(new_requests),
        "paid_requests": len(paid_requests),
        "unpaid_requests": max(len(new_requests) - len(paid_requests), 0),
        "totals": totals,
    }


def generate_excel_24h() -> str:
    events = _last_24h_events()
    wb = Workbook()

    ws1 = wb.active
    border1 = _style_sheet(ws1, "Открытия обмена", ["Время", "Пользователь", "User ID", "Профиль"])
    _add_rows(ws1, _build_opened_rows(events), border1)
    _autosize(ws1)

    ws2 = wb.create_sheet("Заявки")
    border2 = _style_sheet(
        ws2,
        "Заявки",
        ["Время", "Заявка", "Пользователь", "User ID", "Профиль", "Отдаёт", "Получает", "Сумма отдачи", "Сумма получения", "Реквизиты", "Оплата", "Статус", "Оплачено в"],
    )
    _add_rows(ws2, _build_request_rows(events), border2)
    _autosize(ws2)

    ws3 = wb.create_sheet("Сводка")
    border3 = _style_sheet(ws3, "Сводка", ["Показатель", "Значение"])
    fin = _financial_summary(events)
    summary_rows = [
        ["Период", "Последние 24 часа"],
        ["Открыли обмен", fin["opened"]],
        ["Новых заявок", fin["new_requests"]],
        ["Оплачено", fin["paid_requests"]],
        ["Не оплачено", fin["unpaid_requests"]],
    ]
    _add_rows(ws3, summary_rows, border3)
    _autosize(ws3)

    tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.close()
    wb.save(tmp.name)
    return tmp.name


async def _send_text_to_private_chat(text: str) -> bool:
    if not REPORT_BOT_TOKEN or not PRIVATE_CHAT_ID:
        logging.warning("[report_stats] missing REPORT_BOT_TOKEN or PRIVATE_CHAT_ID")
        return False
    url = f"https://api.telegram.org/bot{REPORT_BOT_TOKEN}/sendMessage"
    payload = {"chat_id": str(PRIVATE_CHAT_ID), "text": text, "disable_web_page_preview": True}
    try:
        status, body = await asyncio.to_thread(_post_json, url, payload, 20)
        logging.warning(f"[report_stats] send text status={status} body={body}")
        return True
    except Exception as e:
        logging.exception(f"[report_stats] send text failed: {e}")
        return False


async def _send_document_to_private_chat(file_path: str, caption: str) -> bool:
    if not REPORT_BOT_TOKEN or not PRIVATE_CHAT_ID:
        logging.warning("[report_stats] missing REPORT_BOT_TOKEN or PRIVATE_CHAT_ID")
        return False
    url = f"https://api.telegram.org/bot{REPORT_BOT_TOKEN}/sendDocument"
    try:
        content = Path(file_path).read_bytes()
        status, body = await asyncio.to_thread(
            _post_multipart,
            url,
            {"chat_id": str(PRIVATE_CHAT_ID), "caption": caption},
            "document",
            Path(file_path).name,
            content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            40,
        )
        logging.warning(f"[report_stats] send document status={status} body={body}")
        return True
    except Exception as e:
        logging.exception(f"[report_stats] send document failed: {e}")
        return False


def _read_last_sent_ts() -> str:
    path = Path(STATE_FILE)
    if not path.exists():
        return ""
    try:
        return path.read_text(encoding="utf-8").strip()
    except Exception:
        return ""


def _write_last_sent_ts(value: str) -> None:
    Path(STATE_FILE).write_text(value, encoding="utf-8")


async def _auto_report_loop():
    while True:
        try:
            hours = max(int(AUTO_REPORT_HOURS), 1)
            last_sent = _read_last_sent_ts()
            now = datetime.now(MOSCOW_TZ).replace(tzinfo=None)
            should_send = False

            if not last_sent:
                should_send = True
            else:
                try:
                    dt = datetime.strptime(last_sent, "%Y-%m-%d %H:%M:%S")
                    should_send = (now - dt) >= timedelta(hours=hours)
                except Exception:
                    should_send = True

            if should_send:
                file_path = generate_excel_24h()
                fin = _financial_summary(_last_24h_events())
                from_lines = "\n".join(f"• {cur}: {amt:.8f}" for cur, amt in sorted(fin["totals"].items())) or "—"
                summary_text = (
                    "📊 Отчет за 24ч\n\n"
                    f"👥 Открыли обмен: {fin['opened']}\n"
                    f"🧾 Создано заявок: {fin['new_requests']}\n"
                    f"💰 Оплачено: {fin['paid_requests']}\n"
                    f"⏳ Не оплачено: {fin['unpaid_requests']}\n\n"
                    "💱 Оборот:\n"
                    f"{from_lines}"
                )
                await _send_document_to_private_chat(file_path, "📊 Автоотчёт Excel за 24 часа")
                await _send_text_to_private_chat(summary_text)
                _write_last_sent_ts(_now_str())
        except Exception as e:
            logging.exception(f"[report_stats] auto loop failed: {e}")

        await asyncio.sleep(300)


async def ensure_stats_scheduler_started() -> None:
    global _scheduler_started
    if _scheduler_started:
        return
    _scheduler_started = True
    asyncio.create_task(_auto_report_loop())
