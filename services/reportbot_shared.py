from tempfile import NamedTemporaryFile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def _style_sheet(ws, title: str, headers: list[str]):
    ws.title = title
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 24
    return border


def _autosize(ws):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            val = "" if cell.value is None else str(cell.value)
            widths[cell.column] = max(widths.get(cell.column, 0), min(len(val) + 2, 40))
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _add_rows(ws, rows: list[list], border):
    alt_fill = PatternFill(fill_type="solid", fgColor="F7FBFF")
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if r_idx % 2 == 0:
                cell.fill = alt_fill


def generate_excel_report_24h(storage) -> str:
    opened = storage.get_opened_rows_24h()
    requests = storage.get_request_rows_24h()
    stats = storage.get_stats_24h()

    wb = Workbook()
    ws1 = wb.active
    border1 = _style_sheet(ws1, "Открытия обмена", ["Время", "Пользователь", "User ID", "Профиль"])
    rows1 = [[str(r["created_at"]), r.get("username",""), r.get("user_id",""), r.get("profile_link", "")] for r in opened]
    _add_rows(ws1, rows1, border1)
    _autosize(ws1)

    ws2 = wb.create_sheet("Заявки")
    border2 = _style_sheet(
        ws2,
        "Заявки",
        ["Время", "Заявка", "Пользователь", "User ID", "Профиль", "Отдаёт", "Получает", "Сумма отдачи", "Сумма получения", "Реквизиты", "Оплата", "Статус", "Оплачено в"],
    )
    rows2 = []
    for r in requests:
        rows2.append([
            str(r["created_at"]),
            r["request_id"],
            r.get("username", ""),
            r.get("user_id", ""),
            r.get("profile_link", ""),
            r.get("from_currency", ""),
            r.get("to_currency", ""),
            r.get("amount_from", ""),
            r.get("amount_to", ""),
            r.get("receive_details", ""),
            f"{r.get('payment_method','')} / {r.get('payment_submethod','')}",
            "Оплачено" if r.get("paid_at") else "Ожидает оплаты",
            str(r["paid_at"]) if r.get("paid_at") else "",
        ])
    _add_rows(ws2, rows2, border2)
    _autosize(ws2)

    ws3 = wb.create_sheet("Сводка")
    border3 = _style_sheet(ws3, "Сводка", ["Показатель", "Значение"])
    totals = "\n".join(f"{cur}: {amt:.8f}" for cur, amt in sorted(stats["totals"].items())) or "—"
    rows3 = [
        ["Период", "Последние 24 часа"],
        ["Открыли обмен", stats["opened"]],
        ["Новых заявок", stats["new_requests"]],
        ["Оплачено", stats["paid"]],
        ["Не оплачено", stats["unpaid"]],
        ["QR запросы", stats["qr"]],
        ["Срочных wallet", stats["wallet"]],
        ["Оборот", totals],
    ]
    _add_rows(ws3, rows3, border3)
    _autosize(ws3)

    tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.close()
    wb.save(tmp.name)
    return tmp.name
